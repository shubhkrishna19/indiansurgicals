from __future__ import annotations

import json
import re
import shutil
import unicodedata
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "Company Info"
GENERATED_DIR = ROOT / "src" / "data" / "generated"
OUTPUT_JSON = GENERATED_DIR / "catalogue-source.json"
MEDIA_DIR = ROOT / "public" / "catalogue-media"
HQ_SOURCE_DIR = ROOT / "indian surgicals product images high quality"
HQ_MEDIA_DIR = ROOT / "public" / "catalogue-media-hq"
IMAGE_REPORT_JSON = GENERATED_DIR / "catalogue-image-report.json"
DOWNLOADS_DIR = ROOT / "public" / "downloads"
EXPORT_WORKBOOK = DOWNLOADS_DIR / "indian-surgical-industries-catalogue.xlsx"
EXPORT_PREVIEW_PATH = "/catalogue-preview"
EXPORT_WORKBOOK_PATH = f"/downloads/{EXPORT_WORKBOOK.name}"


CATEGORY_CONFIG = [
    {
        "slug": "autoclaves",
        "name": "Autoclaves",
        "summary": "Portable, vertical, horizontal, dental, semi automatic, and fully automatic sterilization systems.",
        "intro": "Sterilization systems for clinics, hospitals, CSSD workflows, and institutional setups.",
        "order": 1,
    },
    {
        "slug": "sterilizers",
        "name": "Sterilizers",
        "summary": "Instrument and utensil sterilizers in multiple constructions, sizes, and heating formats.",
        "intro": "Dedicated sterilizer ranges for instrument handling, tray lifting, and routine utility sterilization.",
        "order": 2,
    },
    {
        "slug": "hospital-hollowares",
        "name": "Hospital Hollowares",
        "summary": "Dressing drums, trays, kidney trays, bowls, basins, jars, and ward utility stainless steel products.",
        "intro": "Stainless steel holloware for OT, ward, dressing, and general hospital utility use.",
        "order": 3,
    },
    {
        "slug": "suction-units",
        "name": "Suction Units",
        "summary": "Electric, manual, portable, ambulance, pediatric, and foot-operated suction solutions.",
        "intro": "Suction systems and accessories for general procedures, transport, and emergency applications.",
        "order": 4,
    },
    {
        "slug": "needle-destroyers",
        "name": "Needle Destroyers",
        "summary": "Manual and electrical needle and syringe destruction solutions.",
        "intro": "Compact disposal support products for safer sharps handling workflows.",
        "order": 5,
    },
    {
        "slug": "fumigators-foggers",
        "name": "Fumigators and Foggers",
        "summary": "Portable fumigation and fogging equipment for disinfection and controlled application.",
        "intro": "Disinfection support products for hospital and facility hygiene routines.",
        "order": 6,
    },
    {
        "slug": "x-ray-illuminators",
        "name": "X-Ray Illuminators",
        "summary": "Single, double, multi-film, and slim X-ray viewer systems.",
        "intro": "Illumination systems for film viewing and diagnostic interpretation workflows.",
        "order": 7,
    },
    {
        "slug": "hospital-furniture",
        "name": "Hospital Furniture",
        "summary": "ICU beds, Fowler beds, couches, delivery tables, operation tables, and patient furniture.",
        "intro": "Patient care furniture, beds, couches, and procedural support platforms for healthcare facilities.",
        "order": 8,
    },
    {
        "slug": "ward-equipments",
        "name": "Ward Equipments",
        "summary": "Ward utility trolleys, stools, bins, cabinets, cradles, and daily-use support equipment.",
        "intro": "Hospital support equipment for ward organization, bedside utility, and movement inside facilities.",
        "order": 9,
    },
    {
        "slug": "patient-transfer-trolleys",
        "name": "Patient Transfer Trolleys",
        "summary": "Stretchers, emergency recovery trolleys, and transfer solutions.",
        "intro": "Transfer and movement equipment for emergency, recovery, and intra-facility transport.",
        "order": 10,
    },
    {
        "slug": "operation-theater-lights",
        "name": "Operation Theater Lights",
        "summary": "Examination lights and OT light systems in mobile, ceiling, and twin ceiling configurations.",
        "intro": "Lighting systems for examination, OT, and procedure rooms.",
        "order": 11,
    },
]


SECTION_CATEGORY_MAP = {}
for serial in range(1, 24):
    SECTION_CATEGORY_MAP[serial] = "autoclaves"
for serial in range(24, 29):
    SECTION_CATEGORY_MAP[serial] = "sterilizers"
for serial in range(29, 52):
    SECTION_CATEGORY_MAP[serial] = "hospital-hollowares"
for serial in range(52, 62):
    SECTION_CATEGORY_MAP[serial] = "suction-units"
for serial in range(62, 65):
    SECTION_CATEGORY_MAP[serial] = "needle-destroyers"
for serial in range(65, 67):
    SECTION_CATEGORY_MAP[serial] = "fumigators-foggers"
SECTION_CATEGORY_MAP[67] = "x-ray-illuminators"
for serial in range(68, 75):
    SECTION_CATEGORY_MAP[serial] = "hospital-furniture"
for serial in range(75, 83):
    SECTION_CATEGORY_MAP[serial] = "ward-equipments"
for serial in range(83, 85):
    SECTION_CATEGORY_MAP[serial] = "patient-transfer-trolleys"
SECTION_CATEGORY_MAP[85] = "ward-equipments"
for serial in range(86, 90):
    SECTION_CATEGORY_MAP[serial] = "operation-theater-lights"


GENERIC_SECTION_TITLES = {
    "chairs/couchs",
    "chairs and couches",
    "delivery tables",
    "emergency trolleys",
    "hospital bed",
    "hospital furniture",
    "hospital fowler beds",
    "hopital furniture",
    "mattress for hospital beds",
    "operation tables",
    "patient transfer tolleys",
    "patient transfer trolleys",
    "ward equipments",
}


SECTION_NAME_OVERRIDES = {
    "chairs/couchs": "Chairs and Couches",
    "examination light": "Examination Lights",
    "hopital furniture": "Hospital Furniture",
    "o t light ceiling": "OT Light Ceiling",
    "o t light ceiling twin": "OT Light Ceiling Twin",
    "o t light mobile": "OT Light Mobile",
    "patient transfer tolleys": "Patient Transfer Trolleys",
    "x-ray illuminator": "X-Ray Illuminators",
    "x-ray illuminators": "X-Ray Illuminators",
}


MASTER_WORKBOOK = SOURCE_DIR / "catalogue frame.xlsx"
SPEC_WORKBOOKS = [
    SOURCE_DIR / "AUTOCLAVE ALL SPECS  Specification (1).xlsx",
    SOURCE_DIR / "Hospital furniture  specs-2 part (1).xlsx",
]


NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


IMAGE_TOKEN_STOPWORDS = {
    "all",
    "and",
    "body",
    "compact",
    "copy",
    "deluxe",
    "electric",
    "electrical",
    "equipment",
    "film",
    "fix",
    "fixed",
    "fold",
    "folding",
    "for",
    "four",
    "frame",
    "heavy",
    "hospital",
    "hydraulic",
    "india",
    "indian",
    "industries",
    "life",
    "manual",
    "mechanical",
    "metal",
    "mobile",
    "model",
    "new",
    "non",
    "only",
    "part",
    "plain",
    "portable",
    "semi",
    "single",
    "solid",
    "ss",
    "std",
    "surface",
    "table",
    "top",
    "tray",
    "twin",
    "two",
    "type",
    "unit",
    "with",
}


FAMILY_IMAGE_QUERY_RULES = [
    (["wing nut"], ["autoclave wing nut type"]),
    (["p type", "pressure cooker"], ["autoclave p type"]),
    (["digital controller", "controller"], ["autoclave p type digital", "p type with controller"]),
    (["double wall"], ["autoclave double wall"]),
    (["horizontal rectangular"], ["autoclave horizontal rectangular"]),
    (["fully automatic"], ["autoclave fully automatic", "fully automatic"]),
    (["table top", "front loading"], ["table top autoclave"]),
    (["bowl and utensils sterilizer", "bowl sterilizer"], ["bowl sterilizer"]),
    (["instrument sterilizer", "sterilizers"], ["sterilizer"]),
    (["vacuum extractor"], ["suction deluxe", "suction life line ss", "suction ss plain"]),
    (["suction unit"], ["suction", "suction deluxe"]),
    (["suction apparatus"], ["suction", "suction apparatus"]),
    (["slow suction", "pediatric"], ["suction ss plain", "suction"]),
    (["foot operated"], ["suction", "suction life line ss"]),
    (["glass jar", "polycarbonate jar"], ["suction acce", "suction"]),
    (["x ray viewer", "x-ray viewer", "illuminator"], ["x-ray illuminator"]),
    (["icu bed"], ["icu bed"]),
    (["fowler bed"], ["fowler bed"]),
    (["semi fowler"], ["fowler bed", "hospital bed"]),
    (["hospital bed"], ["hospital bed"]),
    (["opd couch chair", "opd couch/chair"], ["opd couch-chair"]),
    (["opd couch"], ["opd couch"]),
    (["pediatric bed"], ["pediatric bed"]),
    (["motorized backrest", "recliner"], ["motorized backrest"]),
    (["mattress"], ["mattress"]),
    (["medicine cupboard", "medicine cupbard"], ["medicine cupboard"]),
    (["cardiac table"], ["cardiac table"]),
    (["mayo", "instrument trolley"], ["mayo instrument trolley", "mayo's instrument trolley"]),
    (["wash basin stand"], ["wash basin stand"]),
    (["i v stand", "iv stand"], ["iv stand"]),
    (["bed side screen"], ["bed side screen"]),
    (["kick bowl", "kick bowel"], ["kick bowel", "kick bowl"]),
    (["recovery trolley"], ["recovery trolley"]),
    (["folding stretcher"], ["folding stretcher"]),
    (["spine board"], ["spine board"]),
    (["ambulance stretcher"], ["ambulance stretcher"]),
    (["stretcher"], ["stretcher"]),
    (["laparoscopic", "monitor trolley"], ["laparoscopic"]),
    (["crash cart"], ["crash cart"]),
    (["medicine trolley"], ["medicine trolley"]),
    (["utility trolley"], ["utility trolley"]),
    (["oxygen cylinder"], ["oxygen cylinder"]),
    (["foot step"], ["foot step"]),
    (["dustbin"], ["dustbin"]),
    (["waste bin"], ["waste bin"]),
    (["biomedical waste trolley"], ["biomedical waste trolley"]),
    (["revolving stool"], ["revolving stool", "patient stool"]),
    (["patient stool"], ["patient stool"]),
    (["surgeon chair"], ["surgeon chair"]),
    (["baby cradle"], ["baby cradle"]),
    (["baby bassinet"], ["baby bassinet"]),
    (["instrument cabinet"], ["instrument cabinet"]),
    (["labour delivery room bed"], ["labour delivery room bed"]),
    (["delivery bed"], ["delivery bed"]),
    (["obstetric"], ["obstetric table"]),
    (["labour table"], ["labour table"]),
    (["gynecological couch", "gynaecological couch"], ["gynecological couch"]),
    (["blood chair", "blood transfusion"], ["blood transfusion chair"]),
    (["dialysis chair"], ["dialysis chair"]),
    (["examination couch"], ["examination couch", "examination table"]),
    (["gynaec examination table", "gynaecological examination"], ["gynaec examination table", "examination table"]),
    (["operation table", "c arm compatible"], ["operation table", "operating table"]),
    (["operating table"], ["operating table", "operation table"]),
    (["spica table"], ["spica table"]),
    (["mortuary trolley"], ["mortuary trolley"]),
    (["autopsy table"], ["autopsy table"]),
    (["bariatric"], ["bariatric table"]),
    (["infant warmer"], ["infant warmer"]),
    (["phototherapy"], ["phototherapy unit", "baby bassinet"]),
    (["ot light"], ["ot light", "pic"]),
    (["examination light", "examination lights"], ["ot light", "pic"]),
    (["bed pan"], ["bedpan"]),
]


CATEGORY_IMAGE_FALLBACKS = {
    "autoclaves": ["autoclave"],
    "sterilizers": ["sterilizer"],
    "hospital-hollowares": ["bedpan"],
    "suction-units": ["suction"],
    "needle-destroyers": [],
    "fumigators-foggers": [],
    "x-ray-illuminators": ["x-ray illuminator"],
    "hospital-furniture": ["ward equipments"],
    "ward-equipments": ["ward equipments", "utility trolley"],
    "patient-transfer-trolleys": ["stretcher", "recovery trolley"],
    "operation-theater-lights": ["ot light", "pic"],
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2022", "-")
    text = text.replace("\u00bd", "1/2").replace("\u00bc", "1/4").replace("\u00be", "3/4")
    text = text.replace("\u00a0", " ")
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text.replace("\n", " ").replace("\r", " ")).strip()


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def slugify(value: str) -> str:
    return re.sub(r"-{2,}", "-", re.sub(r"[^a-z0-9]+", "-", clean_text(value).lower()).strip("-")) or "item"


def humanize_title(value: str) -> str:
    text = clean_text(value)
    if not text:
        return ""
    normalized = normalize_key(text)
    for source, target in SECTION_NAME_OVERRIDES.items():
        if normalize_key(source) == normalized:
            return target
    letters = [char for char in text if char.isalpha()]
    if letters and sum(1 for char in letters if char.isupper()) / len(letters) > 0.55:
        text = text.title()
    replacements = {
        "Abs": "ABS",
        "Ac/Dc": "AC/DC",
        "Cssd": "CSSD",
        "Hp": "HP",
        "Icu": "ICU",
        "Iv": "IV",
        "Led": "LED",
        "Ms": "MS",
        "Opd": "OPD",
        "Ot": "OT",
        "Ss": "SS",
        "X Ray": "X-Ray",
    }
    for source, target in replacements.items():
        text = re.sub(rf"\b{re.escape(source)}\b", target, text)
    return (
        text.replace("Auctoclave", "Autoclave")
        .replace("Hydralic", "Hydraulic")
        .replace("Machanical", "Mechanical")
        .replace("Matterss", "Mattress")
        .replace("Folwer", "Fowler")
        .replace("Couchs", "Couches")
        .replace("Obsteric", "Obstetric")
    )


def unique_preserve(items: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in items:
        cleaned = clean_text(item)
        if not cleaned:
            continue
        key = normalize_key(cleaned)
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def join_compact(items: list[str], limit: int | None = None, separator: str = " | ") -> str:
    values = unique_preserve(items)
    if limit is not None:
        values = values[:limit]
    return separator.join(values)


def format_attribute_pairs(attributes: list[dict[str, str]], limit: int | None = None) -> str:
    pairs = []
    for attribute in attributes:
        label = clean_text(attribute.get("label", ""))
        value = clean_text(attribute.get("value", ""))
        if not label or not value or normalize_key(label) == "model":
            continue
        pairs.append(f"{label}: {value}")
    return join_compact(pairs, limit=limit)


def style_sheet_title(sheet: Any, row: int, title: str, end_column: int = 8) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="194D80")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_sheet_subtitle(sheet: Any, row: int, text: str, end_column: int = 8) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=10, italic=True, color="3F4B59")
    cell.fill = PatternFill("solid", fgColor="EEF3F7")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_table_header(row: Any, fill_color: str = "C96F28") -> None:
    for cell in row:
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_sheet(sheet: Any, min_width: int = 12, max_width: int = 42) -> None:
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        max_length = 0
        for cell in column_cells:
            value = clean_text(cell.value)
            if value:
                max_length = max(max_length, len(value))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, min_width), max_width)


def style_body_cells(sheet: Any, start_row: int, end_row: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_catalogue_workbook(catalogue: dict[str, Any], generated_at: str) -> None:
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    categories_sheet = workbook.create_sheet("Categories")
    families_sheet = workbook.create_sheet("Families")
    models_sheet = workbook.create_sheet("Models")

    family_count = len(catalogue["families"])
    variant_count = sum(len(family["variants"]) for family in catalogue["families"])

    style_sheet_title(overview, 1, "Indian Surgical Industries Public Catalogue", end_column=6)
    style_sheet_subtitle(
        overview,
        2,
        "Structured export arranged for distributors, institutions, marketing circulation, and quick range review. All products are quote only.",
        end_column=6,
    )

    overview.cell(row=4, column=1, value="Generated")
    overview.cell(row=4, column=2, value=generated_at)
    overview.cell(row=5, column=1, value="Categories")
    overview.cell(row=5, column=2, value=len(catalogue["categories"]))
    overview.cell(row=6, column=1, value="Product families")
    overview.cell(row=6, column=2, value=family_count)
    overview.cell(row=7, column=1, value="Model variants")
    overview.cell(row=7, column=2, value=variant_count)
    overview.cell(row=8, column=1, value="Download")
    overview.cell(row=8, column=2, value=EXPORT_WORKBOOK_PATH)
    overview.cell(row=9, column=1, value="Preview")
    overview.cell(row=9, column=2, value=EXPORT_PREVIEW_PATH)

    for cell in overview["A4":"A9"]:
        for item in cell:
            item.font = Font(name="Calibri", bold=True, color="194D80")

    overview.append([])
    overview.append(["Category", "Families", "Model Variants", "Summary"])
    style_table_header(overview[11], fill_color="194D80")
    for category in catalogue["categories"]:
        overview.append(
            [
                category["name"],
                category["familyCount"],
                category["variantCount"],
                category["summary"],
            ]
        )
    style_body_cells(overview, 12, overview.max_row)
    overview.freeze_panes = "A11"
    autosize_sheet(overview, min_width=14, max_width=46)

    style_sheet_title(categories_sheet, 1, "Category Index", end_column=5)
    style_sheet_subtitle(
        categories_sheet,
        2,
        "Concise public-facing category summary for quick buyer navigation and marketing overview.",
        end_column=5,
    )
    categories_sheet.append([])
    categories_sheet.append(["Category", "Families", "Model Variants", "Representative Families", "Category Summary"])
    style_table_header(categories_sheet[4], fill_color="194D80")
    for category in catalogue["categories"]:
        related_families = [family for family in catalogue["families"] if family["categorySlug"] == category["slug"]]
        categories_sheet.append(
            [
                category["name"],
                category["familyCount"],
                category["variantCount"],
                join_compact([family["name"] for family in related_families], limit=5),
                category["summary"],
            ]
        )
    style_body_cells(categories_sheet, 5, categories_sheet.max_row)
    categories_sheet.freeze_panes = "A4"
    autosize_sheet(categories_sheet, min_width=14, max_width=46)

    style_sheet_title(families_sheet, 1, "Product Families", end_column=9)
    style_sheet_subtitle(
        families_sheet,
        2,
        "Family-level export intended for public viewing, advertisement circulation, and shortlist building before quotation discussion.",
        end_column=9,
    )
    families_sheet.append([])
    families_sheet.append(
        [
            "Category",
            "Product Family",
            "Subheading",
            "Models",
            "Summary",
            "Highlights",
            "Representative Models",
            "Image Count",
            "Website Path",
        ]
    )
    style_table_header(families_sheet[4], fill_color="194D80")
    for family in catalogue["families"]:
        families_sheet.append(
            [
                next(
                    (category["name"] for category in catalogue["categories"] if category["slug"] == family["categorySlug"]),
                    family["categorySlug"],
                ),
                family["name"],
                family.get("subheading", ""),
                len(family["variants"]),
                family["summary"],
                join_compact(family.get("highlights", []), limit=3),
                join_compact([variant["model"] for variant in family["variants"]], limit=8),
                len(family.get("images", [])),
                f"/products/{family['categorySlug']}/{family['slug']}",
            ]
        )
    style_body_cells(families_sheet, 5, families_sheet.max_row)
    families_sheet.freeze_panes = "A4"
    autosize_sheet(families_sheet, min_width=14, max_width=48)

    style_sheet_title(models_sheet, 1, "Model Matrix", end_column=7)
    style_sheet_subtitle(
        models_sheet,
        2,
        "Model-level specification export for catalogue comparison, lead qualification, and product understanding without public pricing.",
        end_column=7,
    )
    models_sheet.append([])
    models_sheet.append(
        [
            "Category",
            "Product Family",
            "Model",
            "Key Specifications",
            "Supporting Notes",
            "Image Count",
            "Website Path",
        ]
    )
    style_table_header(models_sheet[4], fill_color="194D80")
    for family in catalogue["families"]:
        category_name = next(
            (category["name"] for category in catalogue["categories"] if category["slug"] == family["categorySlug"]),
            family["categorySlug"],
        )
        for variant in family["variants"]:
            models_sheet.append(
                [
                    category_name,
                    family["name"],
                    variant["model"],
                    format_attribute_pairs(variant.get("attributes", []), limit=8),
                    join_compact(variant.get("descriptionLines", []), limit=3),
                    len(variant.get("images", [])),
                    f"/products/{family['categorySlug']}/{family['slug']}",
                ]
            )
    style_body_cells(models_sheet, 5, models_sheet.max_row)
    models_sheet.freeze_panes = "A4"
    autosize_sheet(models_sheet, min_width=14, max_width=52)

    workbook.save(EXPORT_WORKBOOK)


def tokenize_text(value: str) -> list[str]:
    return [token for token in normalize_key(value).split() if token and token not in IMAGE_TOKEN_STOPWORDS]


def slugify_filename(value: str) -> str:
    return slugify(re.sub(r"\.[A-Za-z0-9]+$", "", clean_text(value)))


def strip_copy_suffix(value: str) -> str:
    text = clean_text(value)
    text = re.sub(r"\bcopy\b(?:\s*\(\d+\))?", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+-\s*$", "", text)
    return clean_text(text)


def build_hq_label(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"^\d+\s*[- ]*", "", stem).strip()
    stem = strip_copy_suffix(stem)
    stem = re.sub(r"\s*\.+\s*$", "", stem)
    return humanize_title(stem)


def infer_category_from_text(text: str) -> str:
    normalized = normalize_key(text)
    if "autoclave" in normalized:
        return "autoclaves"
    if "steriliz" in normalized:
        return "sterilizers"
    if any(
        term in normalized for term in ("bed pan", "bedpan", "kidney tray", "dressing drum", "gallipot", "holloware")
    ):
        return "hospital-hollowares"
    if "suction" in normalized or "vacuum extractor" in normalized:
        return "suction-units"
    if "needle destroyer" in normalized:
        return "needle-destroyers"
    if "fumigator" in normalized or "fogger" in normalized or "disinfector" in normalized:
        return "fumigators-foggers"
    if "x ray" in normalized or "x-ray" in normalized or "illuminator" in normalized:
        return "x-ray-illuminators"
    if "patient transfer" in normalized or "stretcher" in normalized or "spine board" in normalized:
        return "patient-transfer-trolleys"
    if "light" in normalized or "examination" in normalized:
        return "operation-theater-lights"
    if any(
        term in normalized
        for term in ("bed", "delivery table", "operation table", "operating table", "couch", "chair")
    ):
        return "hospital-furniture"
    return "ward-equipments"


def build_family_image_queries(category_slug: str, family_name: str, subheading: str) -> tuple[list[str], bool]:
    normalized = normalize_key(" ".join([family_name, subheading]))
    queries: list[str] = []
    for needles, mapped_queries in FAMILY_IMAGE_QUERY_RULES:
        if any(needle in normalized for needle in needles):
            queries.extend(mapped_queries)
    if queries:
        return unique_preserve(queries), False
    return unique_preserve(CATEGORY_IMAGE_FALLBACKS.get(category_slug, [])), True


def hq_image_matches_query(image: dict[str, Any], query: str) -> bool:
    normalized_query = normalize_key(query)
    if not normalized_query:
        return False
    if normalized_query in image["normalizedLabel"]:
        return True
    query_tokens = normalized_query.split()
    image_tokens = image["tokens"]
    return all(token in image_tokens for token in query_tokens)


def score_hq_image_for_family(
    image: dict[str, Any],
    category_slug: str,
    family_name: str,
    subheading: str,
    queries: list[str],
) -> float:
    normalized_family = normalize_key(" ".join([family_name, subheading]))
    family_tokens = set(tokenize_text(normalized_family))
    score = 0.0

    if image["categorySlug"] == category_slug:
        score += 1.25
    elif category_slug in {"hospital-furniture", "ward-equipments"} and image["categorySlug"] in {
        "hospital-furniture",
        "ward-equipments",
    }:
        score += 0.55
    elif image["categorySlug"] != category_slug:
        score -= 1.0

    if image["normalizedLabel"] == normalize_key(family_name):
        score += 4.0
    elif image["normalizedLabel"] and image["normalizedLabel"] in normalized_family:
        score += 2.4

    overlap = len(family_tokens & image["tokens"])
    if overlap:
        score += min(2.8, overlap * 0.65)

    for query in queries:
        if hq_image_matches_query(image, query):
            score += 1.6

    if "pic" in image["normalizedLabel"] and category_slug == "operation-theater-lights":
        score += 1.2

    if image["serial"] is not None:
        score += 0.02

    return score


def sync_high_quality_media() -> list[dict[str, Any]]:
    if not HQ_SOURCE_DIR.exists():
        return []

    HQ_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    for existing in HQ_MEDIA_DIR.iterdir():
        if existing.is_file():
            existing.unlink()

    image_records: list[dict[str, Any]] = []
    used_targets: set[str] = set()

    for source_path in sorted(HQ_SOURCE_DIR.iterdir()):
        if not source_path.is_file():
            continue

        extension = source_path.suffix.lower()
        if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
            continue

        label = build_hq_label(source_path.name)
        if not label:
            continue
        match = re.match(r"^(\d+)", source_path.stem)
        serial = int(match.group(1)) if match else None

        base_name = slugify_filename(source_path.name)
        target_name = f"{base_name}{extension}"
        duplicate_index = 2
        while target_name in used_targets:
            target_name = f"{base_name}-{duplicate_index}{extension}"
            duplicate_index += 1
        used_targets.add(target_name)

        target_path = HQ_MEDIA_DIR / target_name
        shutil.copyfile(source_path, target_path)

        normalized_label = normalize_key(label)
        image_records.append(
            {
                "originalName": source_path.name,
                "label": label,
                "normalizedLabel": normalized_label,
                "tokens": set(tokenize_text(label)),
                "serial": serial,
                "path": f"/catalogue-media-hq/{target_name}",
                "categorySlug": infer_category_from_text(label),
            }
        )

    return image_records


def collect_hq_family_images(
    category_slug: str,
    family_name: str,
    subheading: str,
    hq_images: list[dict[str, Any]],
    limit: int = 8,
) -> list[str]:
    queries, used_category_fallback = build_family_image_queries(category_slug, family_name, subheading)
    normalized_family = normalize_key(" ".join([family_name, subheading]))
    family_tokens = set(tokenize_text(normalized_family))
    ranked: list[tuple[float, str]] = []

    for image in hq_images:
        query_hits = sum(1 for query in queries if hq_image_matches_query(image, query))
        overlap = len(family_tokens & image["tokens"])
        exact_match = image["normalizedLabel"] == normalize_key(family_name)
        family_contains_label = bool(image["normalizedLabel"] and image["normalizedLabel"] in normalized_family)

        if used_category_fallback:
            if query_hits == 0 and overlap < 2 and not exact_match and not family_contains_label:
                continue
        elif query_hits == 0 and overlap < 2 and not exact_match and not family_contains_label:
            continue

        score = score_hq_image_for_family(image, category_slug, family_name, subheading, queries)
        if query_hits:
            score += min(2.4, query_hits * 0.95)
        if score < 2.15:
            continue
        ranked.append((score, image["path"]))

    ranked.sort(key=lambda item: item[0], reverse=True)
    ordered = [path for _, path in ranked]
    return unique_preserve(ordered[:limit])


def is_page_marker(text: str) -> bool:
    lowered = normalize_key(text)
    return lowered.startswith("page ") or lowered == "prices"


def is_title_candidate(text: str) -> bool:
    lowered = normalize_key(text)
    if not lowered:
        return False
    if lowered in {"model no", "terms"}:
        return False
    if lowered.startswith("optional accessories"):
        return False
    if lowered.startswith("page "):
        return False
    if lowered in {"price list", "net wholesale prices"}:
        return False
    return True


def looks_like_model(text: str) -> bool:
    cleaned = clean_text(text)
    lowered = normalize_key(cleaned)
    if not cleaned or lowered in {"model no", "description", "capacity", "size"}:
        return False
    if lowered.startswith("optional accessories") or lowered.startswith("page "):
        return False
    return bool(re.match(r"^[A-Za-z0-9][A-Za-z0-9 ./-]*\d[A-Za-z0-9 ./-]*$", cleaned))


def extract_catalogue_media() -> list[dict[str, Any]]:
    MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    anchors: list[dict[str, Any]] = []
    with zipfile.ZipFile(MASTER_WORKBOOK) as archive:
        media_files = [name for name in archive.namelist() if name.startswith("xl/media/")]
        for name in media_files:
            target = MEDIA_DIR / Path(name).name
            target.write_bytes(archive.read(name))

        drawing_files = [name for name in archive.namelist() if name.startswith("xl/drawings/drawing") and name.endswith(".xml")]
        for drawing in drawing_files:
            rel_path = drawing.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            rel_targets: dict[str, str] = {}
            if rel_path in archive.namelist():
                rel_root = ET.fromstring(archive.read(rel_path))
                for rel in rel_root:
                    rel_targets[rel.attrib["Id"]] = rel.attrib["Target"].replace("../media/", "")

            root = ET.fromstring(archive.read(drawing))
            for anchor in root:
                from_node = anchor.find("xdr:from", NS)
                blip = anchor.find(".//a:blip", NS)
                if from_node is None or blip is None:
                    continue
                row = int(from_node.find("xdr:row", NS).text) + 1
                relationship_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "")
                media_name = rel_targets.get(relationship_id)
                if not media_name:
                    continue
                anchors.append(
                    {
                        "row": row,
                        "filename": media_name,
                        "path": f"/catalogue-media/{media_name}",
                    }
                )
    anchors.sort(key=lambda item: item["row"])
    return anchors


def detect_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for row_index in range(min(6, len(rows))):
        normalized = [normalize_key(value) for value in rows[row_index]]
        if "product name" in normalized or "product details" in normalized:
            return row_index, normalized
    raise ValueError("Could not detect header row")


def parse_spec_workbooks() -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for workbook_path in SPEC_WORKBOOKS:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            try:
                header_index, normalized_header = detect_header_row(rows)
            except ValueError:
                continue

            header_lookup = {index: normalized_header[index] for index in range(len(normalized_header)) if normalized_header[index]}
            context: dict[str, str] = {}
            current_key = ""

            for row in rows[header_index + 1 :]:
                row_values: dict[str, str] = {}
                for index, header_name in header_lookup.items():
                    if index >= len(row):
                        continue
                    cleaned = clean_text(row[index])
                    if cleaned:
                        row_values[header_name] = cleaned

                if not row_values:
                    continue

                if set(row_values.keys()) == {"description en gb"} and current_key:
                    records[current_key]["descriptionLines"].append(row_values["description en gb"])
                    continue

                model = row_values.get("model", "")
                has_identifier = any(
                    row_values.get(field)
                    for field in ("product name", "model", "product id", "category name", "sub cat")
                )
                if not has_identifier:
                    if current_key and row_values.get("description en gb"):
                        records[current_key]["descriptionLines"].append(row_values["description en gb"])
                    continue

                for field in ("product name", "category name", "sub cat", "shortdescription"):
                    if row_values.get(field):
                        context[field] = row_values[field]
                    elif model and context.get(field):
                        row_values[field] = context[field]

                if not model:
                    if current_key and row_values.get("description en gb"):
                        records[current_key]["descriptionLines"].append(row_values["description en gb"])
                    continue

                key = normalize_key(model)
                current_key = key
                attributes = []
                for field, value in row_values.items():
                    if field in {
                        "product id",
                        "product name",
                        "category name",
                        "sub cat",
                        "model",
                        "shortdescription",
                        "description en gb",
                        "meta title en gb",
                        "meta description en gb",
                        "meta keywords en gb",
                        "image name",
                        "quantity",
                        "price",
                    }:
                        continue
                    attributes.append({"label": humanize_title(field), "value": value})

                existing = records.get(key)
                if existing:
                    existing["descriptionLines"].extend(
                        [row_values.get("description en gb", ""), row_values.get("shortdescription", "")]
                    )
                    existing["attributes"].extend(attributes)
                    existing["sources"].append(
                        {"workbook": workbook_path.name, "sheet": sheet.title, "model": model}
                    )
                    continue

                records[key] = {
                    "model": model,
                    "productName": humanize_title(row_values.get("product name", "")),
                    "categoryName": humanize_title(row_values.get("category name", "")),
                    "subCategory": humanize_title(row_values.get("sub cat", "")),
                    "shortDescription": clean_text(row_values.get("shortdescription", "")),
                    "descriptionLines": [
                        row_values.get("description en gb", ""),
                        row_values.get("shortdescription", ""),
                    ],
                    "attributes": attributes,
                    "sources": [{"workbook": workbook_path.name, "sheet": sheet.title, "model": model}],
                }

        workbook.close()

    for record in records.values():
        record["descriptionLines"] = unique_preserve(record["descriptionLines"])
        deduped: list[dict[str, str]] = []
        seen_pairs: set[tuple[str, str]] = set()
        for attribute in record["attributes"]:
            key = (normalize_key(attribute["label"]), normalize_key(attribute["value"]))
            if key in seen_pairs:
                continue
            seen_pairs.add(key)
            deduped.append(attribute)
        record["attributes"] = deduped

    return records


def build_section_ranges(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    serial_rows: list[tuple[int, int]] = []
    for index, row in enumerate(rows, start=1):
        if len(row) < 2:
            continue
        serial_value = clean_text(row[1])
        if serial_value.isdigit():
            serial_rows.append((int(serial_value), index))

    sections: list[dict[str, Any]] = []
    for position, (serial, row_index) in enumerate(serial_rows):
        next_row_index = serial_rows[position + 1][1] if position + 1 < len(serial_rows) else len(rows) + 1
        sections.append(
            {
                "serial": serial,
                "startRow": row_index,
                "endRow": next_row_index - 1,
            }
        )
    return sections


def find_section_title(rows: list[tuple[Any, ...]], section: dict[str, Any]) -> tuple[str, int]:
    best_score = -1
    best_title = ""
    best_row = section["startRow"]
    for row_index in range(section["startRow"], min(section["startRow"] + 4, section["endRow"] + 1)):
        row = rows[row_index - 1]
        for column_index in (2, 3, 4):
            if column_index >= len(row):
                continue
            candidate = clean_text(row[column_index])
            if not is_title_candidate(candidate):
                continue
            score = len(candidate)
            if row_index == section["startRow"]:
                score += 10
            if column_index == 2:
                score += 4
            if score > best_score:
                best_score = score
                best_title = candidate
                best_row = row_index
    return humanize_title(best_title), best_row


def parse_master_sections(image_anchors: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    workbook = load_workbook(MASTER_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["PRICE LIST"]
    rows = list(sheet.iter_rows(values_only=True))
    sections = build_section_ranges(rows)

    master_variants: dict[str, dict[str, Any]] = {}
    parsed_sections: list[dict[str, Any]] = []

    for section in sections:
        title, title_row = find_section_title(rows, section)
        category_slug = SECTION_CATEGORY_MAP.get(section["serial"], "ward-equipments")
        parsed_section = {
            "serial": section["serial"],
            "title": title,
            "titleRow": title_row,
            "startRow": section["startRow"],
            "endRow": section["endRow"],
            "categorySlug": category_slug,
            "notes": [],
            "models": [],
            "images": [],
        }

        current_header: list[tuple[int, str]] = []
        current_subheading = ""
        current_variant_key = ""

        for row_index in range(section["startRow"], section["endRow"] + 1):
            row = rows[row_index - 1]
            values = [clean_text(cell) for cell in row]
            b_value = values[1] if len(values) > 1 else ""
            c_value = values[2] if len(values) > 2 else ""
            d_value = values[3] if len(values) > 3 else ""
            e_value = values[4] if len(values) > 4 else ""
            f_value = values[5] if len(values) > 5 else ""
            line_values = [value for value in (c_value, d_value, e_value, f_value) if value]

            if not line_values:
                continue
            if row_index == title_row:
                current_subheading = ""
                continue
            if b_value == "TERMS:-" or is_page_marker(b_value):
                continue

            if normalize_key(c_value) == "model no" or any(normalize_key(value) == "model no" for value in line_values):
                current_header = []
                for column_index in range(2, min(len(values), 8)):
                    label = values[column_index]
                    if label:
                        current_header.append((column_index, humanize_title(label)))
                current_variant_key = ""
                continue

            if (
                len(line_values) == 1
                and not looks_like_model(c_value)
                and not normalize_key(c_value).startswith("optional accessories")
                and not normalize_key(c_value).startswith("the above prices")
                and not normalize_key(c_value).startswith("prices models may change")
                and not normalize_key(c_value).startswith("payment terms")
            ):
                current_subheading = humanize_title(c_value)
                continue

            if current_header and c_value and (looks_like_model(c_value) or any((d_value, e_value, f_value))):
                attributes = []
                for column_index, label in current_header[1:]:
                    if column_index >= len(values):
                        continue
                    value = values[column_index]
                    if value:
                        attributes.append({"label": label, "value": value})
                variant = {
                    "model": c_value,
                    "row": row_index,
                    "sectionSerial": section["serial"],
                    "sectionTitle": title,
                    "subheading": current_subheading,
                    "categorySlug": category_slug,
                    "attributes": attributes,
                    "notes": [],
                    "images": [],
                }
                key = normalize_key(c_value)
                master_variants[key] = variant
                parsed_section["models"].append(variant)
                current_variant_key = key
                continue

            if current_variant_key and any((d_value, e_value, f_value)):
                continuation = " ".join(value for value in (c_value, d_value, e_value, f_value) if value)
                if continuation:
                    master_variants[current_variant_key]["notes"].append(continuation)
                continue

            trailing_note = " ".join(line_values)
            if trailing_note:
                parsed_section["notes"].append(trailing_note)

        parsed_sections.append(parsed_section)

    workbook.close()

    for anchor in image_anchors:
        target_section = None
        for section in parsed_sections:
            if section["startRow"] <= anchor["row"] <= section["endRow"]:
                target_section = section
                break
        if not target_section:
            continue

        nearest_variant = None
        nearest_distance = 999
        for variant in target_section["models"]:
            distance = abs(anchor["row"] - variant["row"])
            if distance < nearest_distance:
                nearest_distance = distance
                nearest_variant = variant

        if nearest_variant and nearest_distance <= 4:
            nearest_variant["images"].append(anchor["path"])
        else:
            target_section["images"].append(anchor["path"])

    return parsed_sections, master_variants


def infer_category_from_spec(spec_record: dict[str, Any]) -> str:
    combined = " ".join(
        [
            spec_record.get("categoryName", ""),
            spec_record.get("subCategory", ""),
            spec_record.get("productName", ""),
        ]
    )
    return infer_category_from_text(combined)


def first_meaningful_attribute_value(variant: dict[str, Any]) -> str:
    for attribute in variant.get("attributes", []):
        value = clean_text(attribute.get("value", ""))
        if value:
            return value
    return ""


def section_title_is_generic(section_title: str) -> bool:
    normalized = normalize_key(section_title)
    return normalized in {normalize_key(title) for title in GENERIC_SECTION_TITLES}


def should_use_descriptor_as_family(section_title: str, descriptor: str) -> bool:
    normalized_descriptor = normalize_key(descriptor)
    if not normalized_descriptor:
        return False
    if normalized_descriptor.startswith("isi "):
        return False
    if " litre" in normalized_descriptor or " ltr" in normalized_descriptor:
        return False
    if re.search(r"\b\d+\s*x\s*\d+", normalized_descriptor):
        return False
    if normalized_descriptor in {"ot light", "ot light ceiling", "ot light ceiling twin", "examination light"}:
        return False
    return section_title_is_generic(section_title)


def choose_family_name(section_title: str, subheading: str, descriptor: str, spec_record: dict[str, Any] | None) -> str:
    if spec_record and spec_record.get("productName") and not section_title_is_generic(section_title):
        return humanize_title(spec_record["productName"])
    if should_use_descriptor_as_family(section_title, descriptor):
        return humanize_title(descriptor)
    if spec_record and spec_record.get("productName") and section_title_is_generic(section_title):
        return humanize_title(spec_record["productName"])
    if subheading and section_title_is_generic(section_title) and normalize_key(subheading) != "model no":
        return humanize_title(subheading)
    return humanize_title(section_title)


def merge_attributes(primary: list[dict[str, str]], secondary: list[dict[str, str]]) -> list[dict[str, str]]:
    merged: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for attribute in primary + secondary:
        label = clean_text(attribute.get("label", ""))
        value = clean_text(attribute.get("value", ""))
        if not label or not value:
            continue
        key = (normalize_key(label), normalize_key(value))
        if key in seen:
            continue
        seen.add(key)
        merged.append({"label": humanize_title(label), "value": value})
    return merged


def build_catalogue() -> tuple[dict[str, Any], dict[str, Any]]:
    generated_at = datetime.now().astimezone().replace(microsecond=0).isoformat()
    hq_images = sync_high_quality_media()
    image_anchors = extract_catalogue_media()
    spec_records = parse_spec_workbooks()
    sections, master_variants = parse_master_sections(image_anchors)
    section_lookup = {section["serial"]: section for section in sections}

    families: dict[tuple[str, str], dict[str, Any]] = {}
    processed_models: set[str] = set()

    for model_key, master_variant in master_variants.items():
        spec_record = spec_records.get(model_key)
        category_slug = master_variant["categorySlug"] or infer_category_from_spec(spec_record or {})
        descriptor = first_meaningful_attribute_value(master_variant)
        family_name = choose_family_name(
            master_variant["sectionTitle"],
            master_variant.get("subheading", ""),
            descriptor,
            spec_record,
        )
        family_slug = slugify(family_name)
        family_key = (category_slug, family_slug)

        if family_key not in families:
            section = section_lookup.get(master_variant["sectionSerial"], {})
            families[family_key] = {
                "slug": family_slug,
                "name": family_name,
                "categorySlug": category_slug,
                "sectionSerials": unique_preserve([str(master_variant["sectionSerial"])]),
                "subheading": humanize_title(master_variant.get("subheading", "")),
                "summary": "",
                "description": "",
                "highlights": [],
                "notes": [],
                "images": unique_preserve(section.get("images", [])),
                "variants": [],
                "sourceRefs": [],
            }

        merged_attributes = merge_attributes(
            [{"label": "Model", "value": master_variant["model"]}] + master_variant.get("attributes", []),
            spec_record.get("attributes", []) if spec_record else [],
        )
        description_lines = []
        if spec_record:
            description_lines.extend(spec_record.get("descriptionLines", []))
        description_lines.extend(master_variant.get("notes", []))
        description_lines = unique_preserve(description_lines)

        variant = {
            "model": master_variant["model"],
            "slug": slugify(master_variant["model"]),
            "label": master_variant["model"],
            "attributes": merged_attributes,
            "descriptionLines": description_lines,
            "images": unique_preserve(master_variant.get("images", [])),
            "sourceRefs": [
                {
                    "workbook": MASTER_WORKBOOK.name,
                    "sheet": "PRICE LIST",
                    "sectionSerial": master_variant["sectionSerial"],
                    "row": master_variant["row"],
                }
            ] + (spec_record.get("sources", []) if spec_record else []),
        }

        family = families[family_key]
        family["variants"].append(variant)
        family["images"] = unique_preserve(family["images"] + variant["images"])
        family["notes"] = unique_preserve(family["notes"] + master_variant.get("notes", []))
        if spec_record:
            family["sourceRefs"].extend(spec_record.get("sources", []))
            processed_models.add(model_key)
            if not family["summary"] and spec_record.get("shortDescription"):
                family["summary"] = spec_record["shortDescription"]
            if spec_record.get("descriptionLines") and not family["description"]:
                family["description"] = " ".join(spec_record["descriptionLines"])
            if spec_record.get("descriptionLines"):
                family["highlights"] = unique_preserve(
                    family["highlights"]
                    + [
                        line
                        for line in spec_record["descriptionLines"]
                        if len(line) < 160 and not normalize_key(line).startswith("description")
                    ][:6]
                )

    for model_key, spec_record in spec_records.items():
        if model_key in processed_models:
            continue
        category_slug = infer_category_from_spec(spec_record)
        family_name = humanize_title(spec_record.get("productName") or spec_record.get("subCategory") or spec_record["model"])
        family_slug = slugify(family_name)
        family_key = (category_slug, family_slug)

        if family_key not in families:
            families[family_key] = {
                "slug": family_slug,
                "name": family_name,
                "categorySlug": category_slug,
                "sectionSerials": [],
                "subheading": humanize_title(spec_record.get("subCategory", "")),
                "summary": spec_record.get("shortDescription", ""),
                "description": " ".join(spec_record.get("descriptionLines", [])),
                "highlights": unique_preserve(spec_record.get("descriptionLines", [])[:6]),
                "notes": [],
                "images": [],
                "variants": [],
                "sourceRefs": spec_record.get("sources", []),
            }

        families[family_key]["variants"].append(
            {
                "model": spec_record["model"],
                "slug": slugify(spec_record["model"]),
                "label": spec_record["model"],
                "attributes": merge_attributes(
                    [{"label": "Model", "value": spec_record["model"]}],
                    spec_record.get("attributes", []),
                ),
                "descriptionLines": unique_preserve(spec_record.get("descriptionLines", [])),
                "images": [],
                "sourceRefs": spec_record.get("sources", []),
            }
        )

    family_list = sorted(families.values(), key=lambda item: (item["categorySlug"], item["name"]))
    image_report: list[dict[str, Any]] = []
    for family in family_list:
        family["variants"].sort(key=lambda item: item["model"])
        if not family["summary"]:
            family["summary"] = (
                f"Available in {len(family['variants'])} model option{'s' if len(family['variants']) != 1 else ''}."
            )
        if not family["description"]:
            models = ", ".join(variant["model"] for variant in family["variants"][:6])
            family["description"] = (
                f"{family['name']} is part of the Indian Surgical Industries range and is available in "
                f"{len(family['variants'])} model option{'s' if len(family['variants']) != 1 else ''}."
            )
            if models:
                family["description"] += f" Representative models include {models}."
        if not family["highlights"]:
            family["highlights"] = unique_preserve([family["summary"]])
        labels: list[str] = []
        for variant in family["variants"]:
            for attribute in variant["attributes"]:
                if attribute["label"] == "Model":
                    continue
                if attribute["label"] not in labels:
                    labels.append(attribute["label"])
        family["tableColumns"] = ["Model"] + labels
        hq_family_images = collect_hq_family_images(
            family["categorySlug"],
            family["name"],
            family.get("subheading", ""),
            hq_images,
        )
        family["images"] = unique_preserve(hq_family_images + family["images"])[:10]
        family["notes"] = unique_preserve(family["notes"])
        image_report.append(
            {
                "family": family["name"],
                "categorySlug": family["categorySlug"],
                "matchedHqImages": len(hq_family_images),
                "finalImages": len(family["images"]),
                "images": family["images"][:8],
            }
        )

    categories: list[dict[str, Any]] = []
    for category in CATEGORY_CONFIG:
        related = [family for family in family_list if family["categorySlug"] == category["slug"]]
        preview_images: list[str] = []
        for family in related:
            preview_images.extend(family["images"][:1])
        categories.append(
            {
                **category,
                "familyCount": len(related),
                "variantCount": sum(len(family["variants"]) for family in related),
                "heroImage": preview_images[0] if preview_images else "",
                "gallery": unique_preserve(preview_images[:8]),
            }
        )

    catalogue = {
        "sourceFiles": [MASTER_WORKBOOK.name] + [workbook.name for workbook in SPEC_WORKBOOKS],
        "approvedImageSource": HQ_SOURCE_DIR.name if HQ_SOURCE_DIR.exists() else "",
        "exports": {
            "generatedAt": generated_at,
            "previewPath": EXPORT_PREVIEW_PATH,
            "workbookPath": EXPORT_WORKBOOK_PATH,
            "workbookName": EXPORT_WORKBOOK.name,
            "sheets": ["Overview", "Categories", "Families", "Models"],
        },
        "categories": categories,
        "families": family_list,
        "sections": [
            {
                "serial": section["serial"],
                "title": section["title"],
                "categorySlug": section["categorySlug"],
                "modelCount": len(section["models"]),
            }
            for section in sections
        ],
    }
    report = {
        "importedHighQualityImages": len(hq_images),
        "familiesWithImages": sum(1 for family in family_list if family["images"]),
        "familiesWithoutImages": [family["name"] for family in family_list if not family["images"]],
        "families": image_report,
    }
    return catalogue, report


def main() -> None:
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    catalogue, image_report = build_catalogue()
    write_catalogue_workbook(catalogue, catalogue["exports"]["generatedAt"])
    OUTPUT_JSON.write_text(json.dumps(catalogue, indent=2), encoding="utf-8")
    IMAGE_REPORT_JSON.write_text(json.dumps(image_report, indent=2), encoding="utf-8")
    print(f"Wrote {EXPORT_WORKBOOK}")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {IMAGE_REPORT_JSON}")
    print(f"Categories: {len(catalogue['categories'])}")
    print(f"Families: {len(catalogue['families'])}")


if __name__ == "__main__":
    main()
